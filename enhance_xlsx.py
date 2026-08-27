p = '/home/chinux/jupyterlab/meerkatai/doc_tools.py'
s = open(p).read()

old = '''def md_to_xlsx(md_text, output_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    row_idx = 1
    for typ, payload in _parse_blocks(md_text):
        if typ == "table":
            header, data = _parse_table_rows(payload)
            rows = ([header] if header else []) + data
            for r in rows:
                for ci, cell in enumerate(r, start=1):
                    ws.cell(row=row_idx, column=ci, value=cell)
                row_idx += 1
            row_idx += 1  # 表格之间空一行
        elif typ == "heading":
            level, text = payload
            ws.cell(row=row_idx, column=1, value=text)
            row_idx += 1
        elif typ == "image":
            from openpyxl.drawing.image import Image as XLImage
            _alt, _fmt, b64 = payload
            try:
                img = XLImage(io.BytesIO(base64.b64decode(b64)))
                img.width = 360
                img.height = 360
                ws.add_image(img, f"A{row_idx}")
                row_idx += 22  # 预留图片占用的行
            except Exception:
                ws.cell(row=row_idx, column=1, value=_alt or "[图片]")
                row_idx += 1

    wb.save(output_path)
    return output_path'''

new = '''def md_to_xlsx(md_text, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="top", wrap_text=True)

    row_idx = 1
    for typ, payload in _parse_blocks(md_text):
        if typ == "table":
            header, data = _parse_table_rows(payload)
            rows = ([header] if header else []) + data
            if not rows:
                continue
            start_row = row_idx
            for r in rows:
                for ci, cell in enumerate(r, start=1):
                    ws.cell(row=row_idx, column=ci, value=cell)
                row_idx += 1
            # 表头样式 + 冻结首行
            if header:
                for ci in range(1, len(header) + 1):
                    c = ws.cell(row=start_row, column=ci)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = header_align
                ws.freeze_panes = f"A{start_row + 1}"
            # 列宽自适应
            ncols = max(len(r) for r in rows)
            for ci in range(1, ncols + 1):
                max_len = max((len(str(r[ci - 1])) if ci - 1 < len(r) else 0) for r in rows)
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 50)
            # 数据行对齐
            for ri in range(start_row + 1, row_idx):
                for ci in range(1, ncols + 1):
                    ws.cell(row=ri, column=ci).alignment = cell_align
            row_idx += 1  # 表格之间空一行
        elif typ == "heading":
            level, text = payload
            c = ws.cell(row=row_idx, column=1, value=text)
            c.font = Font(bold=True, size=13 if level <= 2 else 11)
            row_idx += 1
        elif typ == "image":
            from openpyxl.drawing.image import Image as XLImage
            _alt, _fmt, b64 = payload
            try:
                img = XLImage(io.BytesIO(base64.b64decode(b64)))
                img.width = 360
                img.height = 360
                ws.add_image(img, f"A{row_idx}")
                row_idx += 22  # 预留图片占用的行
            except Exception:
                ws.cell(row=row_idx, column=1, value=_alt or "[图片]")
                row_idx += 1

    wb.save(output_path)
    return output_path'''

assert old in s, 'md_to_xlsx 未找到'
s = s.replace(old, new, 1)
open(p, 'w').write(s)
print('md_to_xlsx 增强成功')
